'''Excel-handling file

This file contains the Excel class, which is used to handle excel files.
'''

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory

class Excel:
    """ Clase para manejar archivos excel """
    def __init__(self):
        """ Constructor """
        self.archivo = ''
        self.libro = ''
        self.hoja = ''
        self.celdas = []
        self.datos = []
        self.columnas = []
        self.filas = []

    def crear_archivo(self, nombre):
        """ Crea un archivo excel """
        self.libro = Workbook()
        self.archivo = nombre
        self.hoja = self.libro.active
        self.hoja.title = 'Registros'
        self.celdas = list(self.hoja)
        self.datos = list(self.hoja.values)
        self.columnas = list(self.hoja.columns)
        self.filas = list(self.hoja.rows)
    
    def actualizar_hoja(self):
        """ Actualiza la hoja """
        self.celdas = list(self.hoja)
        self.datos = list(self.hoja.values)
        self.columnas = list(self.hoja.columns)
        self.filas = list(self.hoja.rows)

    def cargar_archivo(self):
        """ Carga el archivo excel """
        Tk().withdraw()
        self.archivo = askopenfilename()
        if self.archivo != '':
            self.libro = load_workbook(self.archivo)
            self.hoja = self.libro.active
            self.celdas = list(self.hoja)
            self.datos = list(self.hoja.values)
            self.columnas = list(self.hoja.columns)
            self.filas = list(self.hoja.rows)
            return True
        else:
            return False

    def guardar_archivo_como(self, nombre=''):
        """ Guarda el archivo excel """
        self.libro.save(nombre if nombre != '' else self.archivo)
    
    def cerrar_archivo(self):
        """ Cierra el archivo excel """
        self.libro.close()
        self.hoja = ''
        self.celdas = []
        self.datos = []
        self.columnas = []
        self.filas = []
        self.archivo = ''
    
    def crear_hoja(self, nombre):
        """ Crea una hoja en el archivo """
        self.hoja = self.libro.create_sheet(nombre)
        self.celdas = list(self.hoja)
        self.datos = list(self.hoja.values)
        self.columnas = list(self.hoja.columns)
        self.filas = list(self.hoja.rows)
    
    def seleccionar_hoja(self, nombre):
        """ Selecciona una hoja del archivo """
        self.hoja = self.libro[nombre]
        self.celdas = list(self.hoja)
        self.datos = list(self.hoja.values)
        self.columnas = list(self.hoja.columns)
        self.filas = list(self.hoja.rows)

    def rellenar_cabeceras(self, cabeceras):
        """ Rellena la primera fila con las cabeceras """
        for i in range(len(cabeceras)):
            self.hoja.cell(row=1, column=i+1).value = cabeceras[i]

    def rellenar_hoja(self, registros):
        """ Rellena la hoja con los registros """
        for registro in registros:
            self.hoja.append(registro)
        self.actualizar_hoja()
    
    def formato_hoja(self):
        """ Formatea la hoja """
        for fila in self.filas[1:]:
            for celda in fila:
                celda.alignment = Alignment(horizontal='center', vertical='center')
                celda.border = Border(
                    left = Side(style='thin'), 
                    right = Side(style='thin'), 
                    top = Side(style='thin'), 
                    bottom = Side(style='thin'))
                celda.font = Font(name='Gotham Light', size=10)
        for columna in self.columnas:
            self.hoja.column_dimensions[columna[0].column_letter].width = 25
    
    def formato_cabecera(self):
        """ Formatea la cabecera """
        for celda in self.celdas[0]:
            celda.alignment = Alignment(horizontal='center', vertical='center')
            celda.border = Border(
                left=Side(style='thick'), 
                right=Side(style='thick'), 
                top=Side(style='thick'), 
                bottom=Side(style='thick'))
            celda.font = Font(name='Gotham Bold', size=12)